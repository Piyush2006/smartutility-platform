from typing import Optional
import csv
import io

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.api.deps import CurrentUser, get_tenant_id, require_permission
from app.core.database import get_db
from app.models.meter import Meter, ReadCycle, RouteMeter
from app.models.reading import ImportRow, MeterReading, MeterReadingImport
from app.schemas.reading import (
    ImportSummaryOut,
    MeterReadingCreate,
    MeterReadingOut,
    RevisitResolve,
    ValidationBreakdownOut,
)
from app.services.audit import record_audit
from app.services.reading_helpers import get_previous_reading
from app.services.storage import save_upload
from app.services.vee_engine import evaluate_reading, resolve_revisit

router = APIRouter(prefix="/meter-readings", tags=["reading"])


def _read_cycle_for_run(db: Session, meter_run_id):
    if meter_run_id is None:
        return None
    from app.models.meter import MeterRun, MeterSchedule

    run = db.get(MeterRun, meter_run_id)
    if run is None:
        return None
    schedule = db.get(MeterSchedule, run.meter_schedule_id)
    return schedule.read_cycle_id if schedule else None


def _create_one_reading(db: Session, tenant_id: str, meter: Meter, current_reading: float, current_reading_date, *, meter_run_id=None, source="manual") -> MeterReading:
    prev_reading, prev_date = get_previous_reading(db, meter.id)
    is_duplicate = prev_date is not None and prev_date == current_reading_date
    reading = MeterReading(
        tenant_id=tenant_id, meter_id=meter.id, meter_run_id=meter_run_id,
        read_cycle_id=_read_cycle_for_run(db, meter_run_id),
        previous_reading=prev_reading, previous_reading_date=prev_date,
        current_reading=current_reading, current_reading_date=current_reading_date,
        source=source, status="Received", is_duplicate=is_duplicate,
    )
    db.add(reading)
    db.commit()
    db.refresh(reading)
    return evaluate_reading(db, reading)


@router.post("", response_model=MeterReadingOut, status_code=status.HTTP_201_CREATED)
def create_reading(
    payload: MeterReadingCreate, tenant_id: str = Depends(get_tenant_id), db: Session = Depends(get_db),
    current: CurrentUser = Depends(require_permission("reading", "reading", "create")),
):
    meter = db.get(Meter, payload.meter_id)
    if meter is None or meter.tenant_id != tenant_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Select a valid meter.")
    reading = _create_one_reading(db, tenant_id, meter, payload.current_reading, payload.current_reading_date, meter_run_id=payload.meter_run_id)
    record_audit(db, tenant_id=tenant_id, user_id=current.id, module="reading", entity="MeterReading", entity_id=reading.id, action="create")
    return reading


@router.get("", response_model=list[MeterReadingOut])
def list_readings(
    tenant_id: str = Depends(get_tenant_id), db: Session = Depends(get_db), status_filter: Optional[str] = None,
    _=Depends(require_permission("reading", "reading", "view")),
):
    stmt = select(MeterReading).where(MeterReading.tenant_id == tenant_id)
    if status_filter:
        stmt = stmt.where(MeterReading.status == status_filter)
    return list(db.execute(stmt).scalars())


@router.get("/template")
def download_template(_=Depends(require_permission("reading", "reading", "download"))):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["device_no", "current_reading", "current_reading_date"])
    writer.writerow(["DEV-0001", "123.456", "2026-01-15"])
    buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=meter_reading_template.csv"})


@router.post("/upload", response_model=ImportSummaryOut, status_code=status.HTTP_201_CREATED)
def upload_readings(
    file: UploadFile = File(...), meter_run_id: Optional[str] = None,
    tenant_id: str = Depends(get_tenant_id), db: Session = Depends(get_db),
    current: CurrentUser = Depends(require_permission("reading", "reading", "create")),
):
    """Upload -> parse -> validate -> create reading records -> VEE. Original
    file + every raw row are kept regardless of validity (CLAUDE.md §15:
    'Never lose original uploaded data')."""
    stored = save_upload(file, sub_dir="reading-uploads", allowed_extensions={".csv", ".xlsx"}, max_mb=10)
    file.file.seek(0)
    raw_bytes = file.file.read()

    rows: list[dict] = []
    if file.filename.lower().endswith(".csv"):
        text = raw_bytes.decode("utf-8-sig")
        for r in csv.DictReader(io.StringIO(text)):
            rows.append(dict(r))
    else:
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2):
            values = [c.value for c in row]
            if all(v is None for v in values):
                continue
            rows.append(dict(zip(headers, values)))

    import_batch = MeterReadingImport(
        tenant_id=tenant_id, meter_run_id=meter_run_id, file_name=file.filename, file_url=stored.url,
        total_rows=len(rows), valid_rows=0, invalid_rows=0,
    )
    db.add(import_batch)
    db.flush()

    valid_count = 0
    invalid_count = 0
    for i, raw in enumerate(rows, start=1):
        error = None
        meter = None
        try:
            device_no = str(raw.get("device_no", "")).strip()
            current_reading = float(raw.get("current_reading"))
            current_reading_date = str(raw.get("current_reading_date")).strip()
            import datetime as _dt
            parsed_date = _dt.date.fromisoformat(current_reading_date[:10])
            if current_reading < 0:
                raise ValueError("current_reading must be >= 0")
            meter = db.execute(select(Meter).where(Meter.tenant_id == tenant_id, Meter.device_no == device_no)).scalar_one_or_none()
            if meter is None:
                raise ValueError(f"Unknown device_no '{device_no}'")
        except Exception as exc:  # noqa: BLE001 -- row-level validation, must never crash the batch
            error = str(exc)

        row_record = ImportRow(tenant_id=tenant_id, import_id=import_batch.id, row_number=i, raw_data=raw, is_valid=error is None, error_message=error)
        db.add(row_record)
        db.flush()

        if error is None:
            reading = _create_one_reading(db, tenant_id, meter, current_reading, parsed_date, meter_run_id=meter_run_id, source="upload")
            row_record.meter_reading_id = reading.id
            valid_count += 1
        else:
            invalid_count += 1

    import_batch.valid_rows = valid_count
    import_batch.invalid_rows = invalid_count
    db.commit()
    db.refresh(import_batch)

    record_audit(db, tenant_id=tenant_id, user_id=current.id, module="reading", entity="MeterReadingImport", entity_id=import_batch.id, action="create", new_value={"valid": valid_count, "invalid": invalid_count})
    return import_batch


@router.post("/{reading_id}/resolve-revisit", response_model=MeterReadingOut)
def resolve_revisit_route(
    reading_id: str, payload: RevisitResolve, tenant_id: str = Depends(get_tenant_id), db: Session = Depends(get_db),
    current: CurrentUser = Depends(require_permission("vee", "vee", "approve")),
):
    reading = db.get(MeterReading, reading_id)
    if reading is None or reading.tenant_id != tenant_id:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Reading not found")
    if reading.status != "Revisit":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only readings in Revisit can be resolved.")
    reading = resolve_revisit(db, reading, corrected_current_reading=payload.corrected_current_reading)
    record_audit(db, tenant_id=tenant_id, user_id=current.id, module="vee", entity="MeterReading", entity_id=reading.id, action="approve", new_value={"status": reading.status})
    return reading


@router.get("/validation-breakdown", response_model=list[ValidationBreakdownOut])
def validation_breakdown(tenant_id: str = Depends(get_tenant_id), db: Session = Depends(get_db), _=Depends(require_permission("vee", "vee", "view"))):
    """Workbook §17 dashboard: per Read Cycle counts across Received/V1/V2/Revisit/Completed."""
    cycles = list(db.execute(select(ReadCycle).where(ReadCycle.tenant_id == tenant_id)).scalars())
    out = []
    for cycle in cycles:
        meter_count = db.execute(select(func.count()).select_from(RouteMeter).where(RouteMeter.route_id == cycle.route_id)).scalar_one()
        readings = list(db.execute(select(MeterReading).where(MeterReading.read_cycle_id == cycle.id)).scalars())
        counts = {"Received": 0, "V1": 0, "V2": 0, "Revisit": 0, "Completed": 0}
        for r in readings:
            counts[r.status] = counts.get(r.status, 0) + 1
        out.append(
            ValidationBreakdownOut(
                read_cycle_id=cycle.id, read_cycle_name=cycle.name, schedule_start_date=None, schedule_end_date=None,
                total_meters=meter_count, readings=len(readings), pending=max(meter_count - len(readings), 0),
                v1=counts["V1"], v2=counts["V2"], revisit=counts["Revisit"], completed=counts["Completed"],
            )
        )
    return out
